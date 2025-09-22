# site_link_compare_frontend.py
import io
import os
import time
import tempfile
from datetime import datetime
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from argparse import Namespace

# Import your existing engine
import Site_Link_compare as slc  # must live in the same folder

# ---------- Helpers ----------
def load_sheet_as_df(xlsx_path: str, sheet: str) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header, *body = rows
    df = pd.DataFrame(body, columns=header)
    return df

def excel_to_bytes(path: str) -> bytes:
    with open(path, "rb") as f:
        return f.read()

def make_default_output_name(prefix: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.xlsx"

def build_args_from_ui(mode: str, out_path: str) -> Namespace:
    # Common toggles
    same_domain_only = st.sidebar.checkbox("Same-domain only", value=True, help="Restrict crawl to the same domain")
    include_subdomains = st.sidebar.checkbox("Include subdomains", value=False)
    keep_query = st.sidebar.checkbox("Keep query parameters", value=False)
    include_fragments = st.sidebar.checkbox("Include #fragment links", value=False)
    pattern_include = st.sidebar.text_input("Pattern include (regex)", value="", help="Only URLs matching this regex are crawled")
    pattern_exclude = st.sidebar.text_input("Pattern exclude (regex)", value="", help="URLs matching this regex are excluded")

    colA, colB = st.sidebar.columns(2)
    max_pages = colA.number_input("Max pages", 1, 5000, 30)
    depth = colB.number_input("Max depth", 0, 10, 2)
    max_links_per_page = colA.number_input("Max links per page", 1, 5000, 300)
    max_total_links = colB.number_input("Max total links", 1, 100000, 3000)
    delay = st.sidebar.number_input("Delay between pages (sec)", 0.0, 10.0, 0.5, 0.1)
    timeout = st.sidebar.number_input("HTTP timeout (sec)", 1, 120, 8)

    st.sidebar.markdown("---")
    respect_robots = st.sidebar.checkbox("Respect robots.txt", value=False)
    headful = st.sidebar.checkbox("Show browser window (headful)", value=False)
    pause_on_first = st.sidebar.checkbox("Pause on first page (for login)", value=False)
    no_resolve = st.sidebar.checkbox("Skip final URL/status resolve (faster)", value=False)

    # Screenshots only relevant for single-site crawl
    if mode == "Single-site crawl":
        no_screenshots = st.sidebar.checkbox("Disable screenshots", value=False)
    else:
        # Compare mode: always disable screenshots (not used)
        no_screenshots = True

    compare_by = "final_url"
    if mode == "Compare sites":
        compare_by = st.sidebar.selectbox(
            "Compare by",
            options=["final_url", "abs_url"],
            index=0,
            help="If you enable 'Skip resolve', prefer 'abs_url' for consistency."
        )
        if no_resolve and compare_by == "final_url":
            # mirror the CLI main() auto-correction
            compare_by = "abs_url"

    args = Namespace(
        # output
        out=out_path,
        # crawl controls
        max_pages=int(max_pages),
        max_links_per_page=int(max_links_per_page),
        max_total_links=int(max_total_links),
        depth=int(depth),
        delay=float(delay),
        same_domain_only=bool(same_domain_only),
        include_subdomains=bool(include_subdomains),
        keep_query=bool(keep_query),
        include_fragments=bool(include_fragments),
        pattern_include=pattern_include,
        pattern_exclude=pattern_exclude,
        # networking / robots
        timeout=int(timeout),
        respect_robots=bool(respect_robots),
        ignore_robots=not bool(respect_robots),
        # browser
        headful=bool(headful),
        pause_on_first=bool(pause_on_first),
        # performance
        no_resolve=bool(no_resolve),
        no_screenshots=bool(no_screenshots),
        # compare mode
        baseline_url=None,
        upgraded_url=None,
        compare_by=compare_by,
        # single crawl
        start_url=None,
    )
    return args

def show_result_tables(out_path: str, mode: str):
    st.success(f"Report created: `{os.path.abspath(out_path)}`")
    if mode == "Single-site crawl":
        sum_df = load_sheet_as_df(out_path, "Crawl Summary")
        links_df = load_sheet_as_df(out_path, "Links")
        if not sum_df.empty:
            st.subheader("Crawl Summary")
            st.dataframe(sum_df, use_container_width=True, hide_index=True)
        if not links_df.empty:
            st.subheader("Links (first 200 rows)")
            st.dataframe(links_df.head(200), use_container_width=True, hide_index=True)
    else:
        sum_df = load_sheet_as_df(out_path, "Summary")
        diff_df = load_sheet_as_df(out_path, "Diff")
        if not sum_df.empty:
            st.subheader("Per-page Summary")
            st.dataframe(sum_df, use_container_width=True, hide_index=True)
        if not diff_df.empty:
            st.subheader("Differences (first 500 rows)")
            st.dataframe(diff_df.head(500), use_container_width=True, hide_index=True)

    st.download_button(
        "⬇️ Download Excel report",
        data=excel_to_bytes(out_path),
        file_name=os.path.basename(out_path),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# ---------- UI ----------
st.set_page_config(page_title="Site Link Crawl & Compare", layout="wide")
st.title("🔗 Site Link Crawl & Compare – Frontend")

st.caption(
    "This UI wraps your engine functions **crawl_and_report** and **compare_upgrade_and_report** "
    "from `Site_Link_compare.py`. Use it to run a crawl or compare two sites, preview results, and download the Excel report."
)

mode = st.radio("Choose mode", ["Single-site crawl", "Compare sites"], horizontal=True)

# Output name and folder
default_name = make_default_output_name("site_links_report" if mode == "Single-site crawl" else "link_diff_report")
col1, col2 = st.columns([3, 2])
out_dir = col1.text_input("Output folder", value=os.path.abspath("."), help="Where to save the Excel report")
out_name = col2.text_input("Filename", value=default_name)

# Build base args
out_path = os.path.join(out_dir, out_name)
args = build_args_from_ui(mode, out_path)

# Mode-specific inputs
if mode == "Single-site crawl":
    start_url = st.text_input("Start URL", placeholder="https://www.example.com")
    args.start_url = start_url.strip() or None
else:
    baseline_url = st.text_input("Baseline site (old)", placeholder="https://old.example.com")
    upgraded_url = st.text_input("Upgraded site (new)", placeholder="https://new.example.com")
    args.baseline_url = baseline_url.strip() or None
    args.upgraded_url = upgraded_url.strip() or None

run_btn = st.button("▶️ Run", type="primary")

# ---------- Run ----------
if run_btn:
    # Validate
    if mode == "Single-site crawl" and not args.start_url:
        st.error("Please enter a Start URL.")
        st.stop()
    if mode == "Compare sites" and (not args.baseline_url or not args.upgraded_url):
        st.error("Please enter both Baseline and Upgraded URLs.")
        st.stop()

    # Make sure output dir exists
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    # Info panel
    with st.status("Running… this may take a while depending on site size and options.", expanded=True) as status:
        st.write("• Launching headless Chrome and fetching pages")
        if args.respect_robots:
            st.write("• Respecting robots.txt (you can disable this in the sidebar)")

        t0 = time.time()
        try:
            if mode == "Single-site crawl":
                # Run your engine
                slc.crawl_and_report(args)  # will print to stdout and save Excel
            else:
                # Ensure screenshots disabled for compare mode
                args.no_screenshots = True
                # Adjust compare_by automatically if needed (mirror CLI main)
                if args.no_resolve and args.compare_by == "final_url":
                    args.compare_by = "abs_url"
                slc.compare_upgrade_and_report(args)
        except Exception as e:
            st.exception(e)
            status.update(label="Failed", state="error")
            st.stop()

        elapsed = time.time() - t0
        st.write(f"• Done in {elapsed:0.1f}s")
        status.update(label="Completed", state="complete", expanded=False)

    # ---------- Show results & download ----------
    show_result_tables(out_path, mode)
