#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Crawl internal pages and generate an Excel report with:
- Cropped screenshots of each <a> element (optional)
- Link text, raw href (absolute), resolved final URL + HTTP status
- Target, rel, internal/external classification
- Source page URL + title for every link

Crawl controls: max pages, depth, delay, same-domain-only, include-subdomains,
robots.txt respect, include/exclude patterns, fragment/query handling.

USAGE - Single-site crawl (original mode):
  python site_link_crawl_report.py \
    --start-url "https://www.example.com" \
    --out "site_links_report.xlsx" \
    --max-pages 40 --depth 2 --delay 0.6 --include-subdomains --respect-robots

USAGE - Compare baseline (old) vs upgraded (new) site (NEW):
  python site_link_crawl_report.py \
    --baseline-url "https://old.example.com" \
    --upgraded-url "https://new.example.com" \
    --out "link_diff_report.xlsx" \
    --depth 2 --max-pages 50 --delay 0.5

Notes for compare mode:
- Discovers pages by crawling the baseline site (BFS by depth/pages).
- For each baseline page, opens the corresponding upgraded page (same path) and extracts links.
- Outputs ONLY differences: Missing, Extra, Wrong (includes target-changed and broken on upgraded).
- Compare key defaults to final_url (post-redirect). Use --no-resolve with --compare-by abs_url for speed.
"""

import argparse
import os
import re
import time
import math
import tempfile
import shutil
from collections import deque
from urllib.parse import urljoin, urlparse, urlunparse, ParseResult
import requests
from PIL import Image
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.common.exceptions import (
    TimeoutException,
    StaleElementReferenceException,
    WebDriverException,
)
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import urllib.robotparser as robotparser

# NEW: typing + dataclasses
from dataclasses import dataclass
from typing import Dict, List, Set, Tuple, Optional

# --------------------------- Utility helpers ---------------------------
def wait_for_ready(driver, timeout=25):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )

def smooth_scroll_to_bottom(driver, step_px=1000, max_wait_s=6):
    """Trigger lazy content with a gentle scroll."""
    try:
        last_height = driver.execute_script("return document.body.scrollHeight")
    except Exception:
        return
    start = time.time()
    while True:
        try:
            driver.execute_script(f"window.scrollBy(0, {step_px});")
        except Exception:
            break
        time.sleep(0.35)
        try:
            new_height = driver.execute_script("return document.body.scrollHeight")
        except Exception:
            break
        if new_height == last_height or (time.time() - start) > max_wait_s:
            break
        last_height = new_height

def sanitize_text(s: str, max_len=250):
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s).strip()
    return s[:max_len]

def is_useful_href(href: str, include_fragments=False):
    if not href:
        return False
    href = href.strip()
    if href.startswith(("javascript:", "mailto:", "tel:", "sms:", "data:")):
        return False
    if not include_fragments and href.startswith("#"):
        return False
    return True

def highlight_element(driver, element):
    try:
        driver.execute_script(
            "arguments[0].setAttribute('data-prev-style', arguments[0].getAttribute('style') || '');"
            "arguments[0].style.outline='3px solid #ff0066';"
            "arguments[0].style.outlineOffset='2px';",
            element,
        )
    except Exception:
        pass

def unhighlight_element(driver, element):
    try:
        driver.execute_script(
            "var prev = arguments[0].getAttribute('data-prev-style');"
            "if (prev !== null) { arguments[0].setAttribute('style', prev); }"
            "arguments[0].removeAttribute('data-prev-style');",
            element,
        )
    except Exception:
        pass

def element_screenshot(element, out_path):
    """Cropped screenshot of the WebElement (Selenium 4)."""
    element.screenshot(out_path)
    return out_path

def resize_image_to_fit_save(in_path, out_path, max_w=380, max_h=140):
    with Image.open(in_path) as im:
        im = im.convert("RGBA")
        w, h = im.size
        scale = min(max_w / w, max_h / h, 1.0)
        if scale < 1.0:
            new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
            im = im.resize(new_size, Image.Resampling.LANCZOS)
        im.save(out_path, format="PNG")
        return im.size  # (w, h) after resize

def px_to_row_height_points(px):
    # Excel row height unit ~ points; approx 1px ≈ 0.75pt
    return max(18, math.ceil(px * 0.75))

def http_resolve(url, timeout=8):
    """Resolve final URL + status (HEAD with GET fallback)."""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; LinkAuditBot/1.0)"}
        r = requests.head(url, allow_redirects=True, timeout=timeout, headers=headers)
        if r.status_code in (405, 403, 400):
            r = requests.get(url, allow_redirects=True, timeout=timeout, headers=headers, stream=True)
        return (r.url, r.status_code)
    except Exception:
        return (url, None)

def excel_set_col_widths(ws, widths):
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_workbook():
    wb = Workbook()
    ws_links = wb.active
    ws_links.title = "Links"
    ws_links.append([
        "Source Page",
        "Source Title",
        "Screenshot",
        "Link Text",
        "Raw Href (absolute)",
        "Resolved URL",
        "HTTP Status",
        "Target",
        "Rel",
        "Internal/External",
    ])
    excel_set_col_widths(ws_links, [50, 40, 55, 45, 80, 80, 12, 12, 18, 16])
    ws_summary = wb.create_sheet("Crawl Summary")
    ws_summary.append(["#", "Page URL", "Page Title", "HTTP Status (fetch)", "Links Captured"])
    excel_set_col_widths(ws_summary, [6, 90, 45, 20, 20])
    return wb, ws_links, ws_summary

def normalize_url(base_url, href, keep_query=False):
    """Return absolute normalized URL (strip fragment; optional query)."""
    abs_url = urljoin(base_url, href)
    parsed = urlparse(abs_url)
    if parsed.scheme not in ("http", "https"):
        return None
    # Remove fragment
    parsed = ParseResult(parsed.scheme, parsed.netloc, parsed.path, parsed.params,
                         parsed.query if keep_query else "", "")
    return urlunparse(parsed)

def host_core(host: str):
    """Remove leading 'www.' for a broader same-domain match (approx)."""
    host = (host or "").lower()
    return host[4:] if host.startswith("www.") else host

def is_internal_url(url: str, base_host: str, include_subdomains=True):
    """Heuristic internal check without extra deps (treat base host and its subdomains as internal)."""
    try:
        netloc = urlparse(url).netloc.lower()
    except Exception:
        return False
    base_core = host_core(base_host)
    netloc_core = host_core(netloc)
    if netloc_core == base_core:
        return True
    if include_subdomains and netloc_core.endswith("." + base_core):
        return True
    return False

def allowed_by_robots(robots, url, user_agent="*"):
    try:
        return robots.can_fetch(user_agent, url)
    except Exception:
        return True

# --------------------------- NEW: Driver factory ---------------------------
def make_driver(args):
    chrome_opts = ChromeOptions()
    if not args.headful:
        chrome_opts.add_argument("--headless=new")
    chrome_opts.add_argument("--window-size=1400,2000")
    chrome_opts.add_argument("--disable-gpu")
    chrome_opts.add_argument("--no-sandbox")
    chrome_opts.add_argument("--disable-dev-shm-usage")
    chrome_opts.add_argument("--disable-blink-features=AutomationControlled")
    chrome_opts.add_argument("--force-device-scale-factor=1")
    chrome_opts.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    )
    return webdriver.Chrome(options=chrome_opts)

# --------------------------- NEW: Data classes ---------------------------
@dataclass
class LinkRec:
    link_text: str
    abs_url: str
    final_url: str
    status: Optional[int]
    target: str
    rel: str

@dataclass
class PageLinks:
    page_url: str
    page_title: str
    fetch_status: Optional[int]
    links: List[LinkRec]

def link_compare_key(rec: LinkRec, compare_by: str = "final_url") -> str:
    # compare_by: "final_url" or "abs_url"
    if compare_by == "abs_url":
        return rec.abs_url or ""
    return (rec.final_url or rec.abs_url or "")

# --------------------------- NEW: Diff workbook ---------------------------
def create_diff_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Diff"
    ws.append([
        "Page URL (baseline)",
        "Page URL (upgraded)",
        "Type",  # Missing | Extra | Wrong
        "Link Text",
        "Baseline URL",
        "Upgraded URL",
        "Baseline Status",
        "Upgraded Status",
        "Note",
    ])
    excel_set_col_widths(ws, [50, 50, 12, 40, 80, 80, 14, 14, 40])
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append([
        "#", "Page URL (baseline)", "Page URL (upgraded)",
        "Baseline Title", "Upgraded Title",
        "Missing", "Extra", "Wrong"
    ])
    excel_set_col_widths(ws_sum, [6, 60, 60, 40, 40, 10, 10, 10])
    return wb, ws, ws_sum

# --------------------------- NEW: Page link extraction (no Excel) ---------------------------
def extract_links_on_page(driver, page_url, args, base_host) -> PageLinks:
    page_title = ""
    fetch_status = None
    try:
        driver.get(page_url)
        wait_for_ready(driver, timeout=35)
        smooth_scroll_to_bottom(driver, step_px=1200, max_wait_s=7)
        page_title = sanitize_text(driver.title, 300)
        if not args.no_resolve:
            _, fetch_status = http_resolve(driver.current_url, timeout=args.timeout)
    except Exception:
        pass

    try:
        anchors = driver.find_elements(By.TAG_NAME, "a")
    except Exception:
        anchors = []

    links: List[LinkRec] = []
    for a in anchors[: max(0, args.max_links_per_page)]:
        try:
            href = a.get_attribute("href")
        except Exception:
            continue
        if not is_useful_href(href, include_fragments=args.include_fragments):
            continue

        abs_url = normalize_url(driver.current_url, href, keep_query=args.keep_query)
        if not abs_url:
            continue

        link_text = ""
        try:
            link_text = a.text or a.get_attribute("aria-label") or ""
            link_text = sanitize_text(link_text, 250)
        except Exception:
            pass

        target = ""
        rel = ""
        try:
            target = a.get_attribute("target") or ""
            rel = a.get_attribute("rel") or ""
        except Exception:
            pass

        final_url, status = (abs_url, None)
        if not args.no_resolve:
            final_url, status = http_resolve(abs_url, timeout=args.timeout)

        links.append(
            LinkRec(
                link_text=link_text,
                abs_url=abs_url,
                final_url=final_url,
                status=status,
                target=target,
                rel=rel,
            )
        )
    return PageLinks(page_url=page_url, page_title=page_title, fetch_status=fetch_status, links=links)

# --------------------------- NEW: Compare logic ---------------------------
def compare_page_links(
    base: PageLinks,
    upg: PageLinks,
    compare_by: str = "final_url"
) -> Tuple[List[List], Dict[str, int]]:
    """
    Returns:
      rows_for_diff_sheet: list of Excel rows for "Diff" sheet
      counts: {"Missing": n, "Extra": n, "Wrong": n}
    'Wrong' includes target-changed for same text and broken on upgraded.
    """
    rows: List[List] = []
    counts = {"Missing": 0, "Extra": 0, "Wrong": 0}

    # Index by compare key
    base_by_key: Dict[str, LinkRec] = {link_compare_key(r, compare_by): r for r in base.links}
    upg_by_key: Dict[str, LinkRec]  = {link_compare_key(r, compare_by): r for r in upg.links}

    base_keys: Set[str] = set(base_by_key.keys())
    upg_keys:  Set[str] = set(upg_by_key.keys())

    # Map of text -> set(keys) for target-changed detection
    def text_map(links: List[LinkRec]) -> Dict[str, Set[str]]:
        d: Dict[str, Set[str]] = {}
        for r in links:
            k = link_compare_key(r, compare_by)
            t = r.link_text.strip()
            if not t:
                # Skip empty-text anchors in text-based 'wrong' rule
                continue
            d.setdefault(t, set()).add(k)
        return d

    base_text_map = text_map(base.links)
    upg_text_map  = text_map(upg.links)

    # Detect target-changed for same anchor text
    texts_common = set(base_text_map.keys()) & set(upg_text_map.keys())
    keys_in_wrong_text: Set[str] = set()
    for t in sorted(texts_common):
        if base_text_map[t] != upg_text_map[t]:
            base_targets = [ (base_by_key[k].final_url if compare_by=="final_url" else base_by_key[k].abs_url)
                             for k in base_text_map[t] if k in base_by_key ]
            upg_targets  = [ (upg_by_key[k].final_url if compare_by=="final_url" else upg_by_key[k].abs_url)
                             for k in upg_text_map[t] if k in upg_by_key ]
            rows.append([
                base.page_url, upg.page_url, "Wrong", t,
                " | ".join(sorted(set(base_targets))),
                " | ".join(sorted(set(upg_targets))),
                "", "", "Target changed for same anchor text"
            ])
            counts["Wrong"] += 1
            keys_in_wrong_text |= base_text_map[t] | upg_text_map[t]

    # Clean sets before missing/extra accounting
    base_keys_clean = base_keys - keys_in_wrong_text
    upg_keys_clean  = upg_keys  - keys_in_wrong_text

    # 'Wrong' due to broken on upgraded side (same key on both)
    inter_keys = base_keys_clean & upg_keys_clean
    for k in sorted(inter_keys):
        b = base_by_key[k]
        u = upg_by_key[k]
        b_status = b.status if b.status is not None else ""
        u_status = u.status if u.status is not None else ""
        # Treat None or >=400 on upgraded as wrong (broken)
        if u.status is None or (isinstance(u.status, int) and u.status >= 400):
            note = "Upgraded link broken"
            rows.append([
                base.page_url, upg.page_url, "Wrong", u.link_text,
                (b.final_url if compare_by=="final_url" else b.abs_url),
                (u.final_url if compare_by=="final_url" else u.abs_url),
                b_status, u_status, note
            ])
            counts["Wrong"] += 1

    # Missing & Extra after removing 'wrong by text' keys
    missing = sorted(base_keys_clean - upg_keys_clean)
    extra   = sorted(upg_keys_clean - base_keys_clean)

    for k in missing:
        b = base_by_key[k]
        rows.append([
            base.page_url, upg.page_url, "Missing", b.link_text,
            (b.final_url if compare_by=="final_url" else b.abs_url),
            "", b.status if b.status is not None else "", "", ""
        ])
        counts["Missing"] += 1

    for k in extra:
        u = upg_by_key[k]
        rows.append([
            base.page_url, upg.page_url, "Extra", u.link_text,
            "", (u.final_url if compare_by=="final_url" else u.abs_url),
            "", u.status if u.status is not None else "", ""
        ])
        counts["Extra"] += 1

    return rows, counts

# --------------------------- NEW: Compare workflow ---------------------------
def compare_upgrade_and_report(args):
    base_origin = urlparse(args.baseline_url)
    upg_origin  = urlparse(args.upgraded_url)

    # robots.txt
    robots_base = robotparser.RobotFileParser()
    robots_upg  = robotparser.RobotFileParser()
    robots_base_read = robots_upg_read = False
    if args.respect_robots:
        try:
            robots_base.set_url(f"{base_origin.scheme}://{base_origin.netloc}/robots.txt")
            robots_base.read()
            robots_base_read = True
        except Exception:
            robots_base_read = False
        try:
            robots_upg.set_url(f"{upg_origin.scheme}://{upg_origin.netloc}/robots.txt")
            robots_upg.read()
            robots_upg_read = True
        except Exception:
            robots_upg_read = False

    # Driver
    try:
        driver = make_driver(args)
    except WebDriverException:
        print("Failed to start Chrome WebDriver. Ensure Chrome is installed and Selenium >= 4.6.")
        raise

    # Workbook for diffs
    wb, ws_diff, ws_summary = create_diff_workbook()

    # BFS over baseline site pages to discover page set
    q = deque([(args.baseline_url, 0)])
    visited: Set[str] = set()
    pages_crawled = 0
    page_idx = 0

    while q and pages_crawled < args.max_pages:
        base_page_url, depth = q.popleft()

        # robots baseline
        if args.respect_robots and robots_base_read and not allowed_by_robots(robots_base, base_page_url):
            continue

        norm_page = normalize_url(base_page_url, "", keep_query=args.keep_query)
        if not norm_page or norm_page in visited:
            continue
        visited.add(norm_page)
        pages_crawled += 1
        page_idx += 1

        # Extract baseline page links
        base_page = extract_links_on_page(driver, base_page_url, args, base_origin.netloc)

        # Build upgraded page URL with same path/params (query optional)
        parsed = urlparse(base_page.page_url)
        upg_page_parsed = ParseResult(
            scheme=upg_origin.scheme,
            netloc=upg_origin.netloc,
            path=parsed.path,
            params=parsed.params,
            query=parsed.query if args.keep_query else "",
            fragment=""
        )
        upg_page_url = urlunparse(upg_page_parsed)

        # robots upgraded
        if args.respect_robots and robots_upg_read and not allowed_by_robots(robots_upg, upg_page_url):
            upg_page = PageLinks(page_url=upg_page_url, page_title="", fetch_status=None, links=[])
        else:
            upg_page = extract_links_on_page(driver, upg_page_url, args, upg_origin.netloc)

        # Compare
        rows, counts = compare_page_links(base_page, upg_page, compare_by=args.compare_by)

        # Write diff rows
        for r in rows:
            ws_diff.append(r)

        # Summary per page
        ws_summary.append([
            page_idx, base_page.page_url, upg_page.page_url,
            base_page.page_title, upg_page.page_title,
            counts["Missing"], counts["Extra"], counts["Wrong"]
        ])

        # Discover further baseline pages (internal + patterns) for BFS
        if depth < args.depth:
            base_host = base_origin.netloc
            discovered: Set[str] = set()
            for r in base_page.links:
                consider = True
                if args.same_domain_only and not is_internal_url(
                    r.abs_url, base_host, include_subdomains=args.include_subdomains
                ):
                    consider = False
                if args.pattern_include and not re.search(args.pattern_include, r.abs_url, flags=re.IGNORECASE):
                    consider = False
                if args.pattern_exclude and re.search(args.pattern_exclude, r.abs_url, flags=re.IGNORECASE):
                    consider = False
                if consider:
                    discovered.add(r.abs_url)
            for u in discovered:
                if args.respect_robots and robots_base_read and not allowed_by_robots(robots_base, u):
                    continue
                nu = normalize_url(u, "", keep_query=args.keep_query)
                if nu and nu not in visited:
                    q.append((u, depth + 1))

        # Politeness
        time.sleep(max(0.0, args.delay))

    # Save results
    wb.save(args.out)
    print(f"✅ Compare complete: {pages_crawled} baseline page(s) assessed")
    print(f"📄 Diff report saved: {os.path.abspath(args.out)}")

    driver.quit()

# --------------------------- Original single-site crawl ---------------------------
def crawl_and_report(args):
    start_url = args.start_url
    base_host = urlparse(start_url).netloc

    # Prepare Selenium Chrome
    chrome_opts = ChromeOptions()
    if not args.headful:
        chrome_opts.add_argument("--headless=new")
    chrome_opts.add_argument("--window-size=1400,2000")
    chrome_opts.add_argument("--disable-gpu")
    chrome_opts.add_argument("--no-sandbox")
    chrome_opts.add_argument("--disable-dev-shm-usage")
    chrome_opts.add_argument("--disable-blink-features=AutomationControlled")
    chrome_opts.add_argument("--force-device-scale-factor=1")
    chrome_opts.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    )
    try:
        driver = webdriver.Chrome(options=chrome_opts)
    except WebDriverException:
        print("Failed to start Chrome WebDriver. Ensure Chrome is installed and Selenium >= 4.6.")
        raise

    # robots.txt
    robots = robotparser.RobotFileParser()
    robots_read = False
    if args.respect_robots:
        try:
            origin = urlparse(start_url)
            robots_url = f"{origin.scheme}://{origin.netloc}/robots.txt"
            robots.set_url(robots_url)
            robots.read()
            robots_read = True
        except Exception:
            robots_read = False

    temp_dir = tempfile.mkdtemp(prefix="link_crawl_screens_")

    # Excel workbook
    wb, ws_links, ws_summary = create_workbook()

    # BFS queue and visited
    q = deque([(start_url, 0)])
    visited = set()
    pages_crawled = 0
    total_links_captured = 0
    link_img_max_w, link_img_max_h = 380, 140

    if args.pause_on_first:
        # Open the first page headfully, so user can login, then continue
        if not args.headful:
            print("NOTE: --pause-on-first is most useful with --headful")
        print("Opening the start URL to let you login/solve CAPTCHA if needed…")
        driver.get(start_url)
        try:
            wait_for_ready(driver, timeout=40)
            input("Press ENTER to start the crawl…")
        except Exception:
            pass
        # Put it back on the queue for normal processing
        q = deque([(driver.current_url, 0)])

    while q and pages_crawled < args.max_pages:
        page_url, depth = q.popleft()

        # Respect robots
        if args.respect_robots and robots_read:
            if not allowed_by_robots(robots, page_url):
                # Skip page disallowed by robots.txt
                continue

        # Normalize page URL for visited set
        normalized_page = normalize_url(page_url, "", keep_query=args.keep_query)
        if not normalized_page:
            continue
        if normalized_page in visited:
            continue
        visited.add(normalized_page)
        pages_crawled += 1

        # Fetch page
        try:
            driver.get(page_url)
            wait_for_ready(driver, timeout=35)
            smooth_scroll_to_bottom(driver, step_px=1200, max_wait_s=7)
            page_title = sanitize_text(driver.title, 300)
            # Infer fetch status via a lightweight HEAD (optional)
            fetch_status = None
            if not args.no_resolve:
                _, fetch_status = http_resolve(driver.current_url, timeout=args.timeout)
        except TimeoutException:
            page_title = ""
            fetch_status = None
        except Exception:
            page_title = ""
            fetch_status = None

        # Collect links on this page
        try:
            anchors = driver.find_elements(By.TAG_NAME, "a")
        except Exception:
            anchors = []

        # First, collect link data and also discover next URLs for BFS
        rows_for_page = []
        discovered_urls = set()
        for idx, a in enumerate(anchors):
            # Skip stale/hidden/zero-size quickly
            try:
                href = a.get_attribute("href")
            except StaleElementReferenceException:
                continue
            except Exception:
                continue
            if not is_useful_href(href, include_fragments=args.include_fragments):
                continue

            abs_url = normalize_url(driver.current_url, href, keep_query=args.keep_query)
            if not abs_url:
                continue

            # BFS discovery (only from links that look internal or allowed by filters)
            consider_for_crawl = True
            if args.same_domain_only and not is_internal_url(abs_url, base_host, include_subdomains=args.include_subdomains):
                consider_for_crawl = False
            if args.pattern_include and not re.search(args.pattern_include, abs_url, flags=re.IGNORECASE):
                consider_for_crawl = False
            if args.pattern_exclude and re.search(args.pattern_exclude, abs_url, flags=re.IGNORECASE):
                consider_for_crawl = False
            if consider_for_crawl and depth < args.depth:
                discovered_urls.add(abs_url)

            # Metadata for Excel
            try:
                link_text = a.text or a.get_attribute("aria-label") or ""
                link_text = sanitize_text(link_text, max_len=250)
                target = a.get_attribute("target") or ""
                rel = a.get_attribute("rel") or ""

                # Prepare screenshot if enabled
                shot_path = ""
                resized_path = ""
                row_height_px = 60
                if not args.no_screenshots:
                    # Only screenshot visible elements with some size
                    try:
                        if a.is_displayed():
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", a)
                            time.sleep(0.12)
                            highlight_element(driver, a)
                            time.sleep(0.05)
                            shot_path = os.path.join(temp_dir, f"p{pages_crawled}_link_{idx+1}.png")
                            element_screenshot(a, shot_path)
                            unhighlight_element(driver, a)
                            # Resize to fit the Excel cell nicely
                            resized_path = os.path.join(temp_dir, f"p{pages_crawled}_link_{idx+1}_res.png")
                            (rw, rh) = resize_image_to_fit_save(
                                shot_path, resized_path, max_w=link_img_max_w, max_h=link_img_max_h
                            )
                            row_height_px = rh if rh else 60
                        else:
                            # not displayed; skip screenshot
                            resized_path = ""
                    except Exception:
                        resized_path = ""

                # Resolve final URL / status
                final_url, status = (abs_url, None)
                if not args.no_resolve:
                    final_url, status = http_resolve(abs_url, timeout=args.timeout)

                internal_external = "Internal" if is_internal_url(final_url or abs_url, base_host, include_subdomains=args.include_subdomains) else "External"

                rows_for_page.append({
                    "source_url": driver.current_url,
                    "source_title": page_title,
                    "img_path": resized_path,  # may be empty
                    "link_text": link_text,
                    "abs_url": abs_url,
                    "final_url": final_url,
                    "status": status,
                    "target": target,
                    "rel": rel,
                    "scope": internal_external,
                    "row_height_px": row_height_px
                })
                if len(rows_for_page) >= args.max_links_per_page:
                    break
            except StaleElementReferenceException:
                continue
            except Exception:
                continue

        # Add discovered URLs to queue
        for u in discovered_urls:
            if args.respect_robots and robots_read and not allowed_by_robots(robots, u):
                continue
            norm_u = normalize_url(u, "", keep_query=args.keep_query)
            if norm_u and norm_u not in visited:
                q.append((u, depth + 1))

        # Write summary row
        ws_summary.append([pages_crawled, driver.current_url, page_title, fetch_status, len(rows_for_page)])

        # Write link rows to Excel
        start_row = ws_links.max_row + 1
        for i, rec in enumerate(rows_for_page, start=0):
            row = start_row + i
            # Text columns
            ws_links.cell(row=row, column=1, value=rec["source_url"])
            ws_links.cell(row=row, column=2, value=rec["source_title"])
            # Image anchor in column C (if available)
            if rec["img_path"]:
                try:
                    xl_img = XLImage(rec["img_path"])
                    ws_links.add_image(xl_img, f"C{row}")
                    ws_links.row_dimensions[row].height = px_to_row_height_points(rec["row_height_px"])
                except Exception:
                    ws_links.row_dimensions[row].height = 60
            else:
                ws_links.row_dimensions[row].height = 60
            ws_links.cell(row=row, column=4, value=rec["link_text"])
            ws_links.cell(row=row, column=5, value=rec["abs_url"])
            ws_links.cell(row=row, column=6, value=rec["final_url"])
            ws_links.cell(row=row, column=7, value=rec["status"] if rec["status"] is not None else "")
            ws_links.cell(row=row, column=8, value=rec["target"])
            ws_links.cell(row=row, column=9, value=rec["rel"])
            ws_links.cell(row=row, column=10, value=rec["scope"])

        total_links_captured += len(rows_for_page)

        # Politeness delay
        time.sleep(max(0.0, args.delay))
        if total_links_captured >= args.max_total_links:
            break

    # Save workbook
    wb.save(args.out)
    print(f"✅ Crawl complete: {pages_crawled} page(s) visited")
    print(f"✅ Links captured: {total_links_captured}")
    print(f"📄 Report saved: {os.path.abspath(args.out)}")

    # Cleanup
    driver.quit()
    shutil.rmtree(temp_dir, ignore_errors=True)

# --------------------------- CLI args ---------------------------
def parse_args():
    p = argparse.ArgumentParser(description="Crawl site pages and export link report to Excel with screenshots, or compare baseline vs upgraded site and export only differences.")
    # Original single-site crawl
    p.add_argument("--start-url", help="Start URL for the crawl")
    p.add_argument("--out", default="site_links_report.xlsx", help="Output Excel path")
    # Crawl controls
    p.add_argument("--max-pages", type=int, default=30, help="Max number of pages to crawl")
    p.add_argument("--max-links-per-page", type=int, default=300, help="Max links per page to capture")
    p.add_argument("--max-total-links", type=int, default=3000, help="Global max links to capture")
    p.add_argument("--depth", type=int, default=2, help="Max crawl depth from start URL")
    p.add_argument("--delay", type=float, default=0.5, help="Delay (seconds) between page fetches")
    p.add_argument("--same-domain-only", action="store_true", default=True,
                   help="Restrict crawl to same domain (default: on)")
    p.add_argument("--include-subdomains", action="store_true", help="Treat subdomains as internal")
    p.add_argument("--keep-query", action="store_true", help="Do not strip query params during normalization")
    p.add_argument("--include-fragments", action="store_true", help="Include #fragment-only links on pages")
    p.add_argument("--pattern-include", default="", help="Regex; only URLs matching are crawled")
    p.add_argument("--pattern-exclude", default="", help="Regex; URLs matching are excluded from crawl and capture")
    # Networking / robots
    p.add_argument("--timeout", type=int, default=8, help="HTTP timeout seconds for resolution")
    group_robots = p.add_mutually_exclusive_group()
    group_robots.add_argument("--respect-robots", action="store_true", help="Obey robots.txt (recommended)")
    group_robots.add_argument("--ignore-robots", action="store_true", help="Ignore robots.txt (use responsibly)")
    # Browser
    p.add_argument("--headful", action="store_true", help="Run a visible browser")
    p.add_argument("--pause-on-first", action="store_true", help="Open first page and wait for ENTER (login etc.)")
    # Output speed toggles
    p.add_argument("--no-resolve", action="store_true", help="Skip resolving final URLs / status (faster)")
    p.add_argument("--no-screenshots", action="store_true", help="Do not capture element screenshots")

    # NEW: Compare mode
    p.add_argument("--baseline-url", help="Baseline site start URL (old site)")
    p.add_argument("--upgraded-url", help="Upgraded site base URL (new site)")
    p.add_argument(
        "--compare-by",
        choices=["final_url", "abs_url"],
        default="final_url",
        help="Compare links using final_url (default) or abs_url (use abs when --no-resolve)."
    )

    args = p.parse_args()

    # Validate mode selection
    if not args.start_url and not (args.baseline_url and args.upgraded_url):
        p.error("Provide either --start-url (single crawl) OR (--baseline-url AND --upgraded-url) for compare mode.")

    return args

# --------------------------- Main ---------------------------
if __name__ == "__main__":
    args = parse_args()
    # Compare mode if both provided
    if args.baseline_url and args.upgraded_url:
        # Prefer to disable screenshots in compare mode (no need for element crops)
        if not getattr(args, "no_screenshots", False):
            args.no_screenshots = True
        # If no_resolve is set but user left default compare_by, switch to abs_url for consistent comparison
        if args.no_resolve and args.compare_by == "final_url":
            args.compare_by = "abs_url"
        compare_upgrade_and_report(args)
    else:
        # Fall back to original single-site crawl
        crawl_and_report(args)
