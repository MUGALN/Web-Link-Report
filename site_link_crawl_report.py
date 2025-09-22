#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Crawl internal pages and generate an Excel report with:
- Cropped screenshots of each <a> element (optional)
- Link text, raw href (absolute), resolved final URL + HTTP status
- Target, rel, internal/external classification
- Source page URL + title for every link
Crawl controls: max pages, depth, delay, same-domain-only, include-subdomains,
robots.txt respect, include/exclude patterns, fragment/query handling.

Usage:
  python site_link_crawl_report.py \
    --start-url "https://example.com" \
    --out "site_links_report.xlsx" \
    --max-pages 40 --depth 2 --delay 0.6 --include-subdomains --respect-robots
"""

import argparse
import os
import re
import time
import math
import tempfile
import shutil
from collections import deque
from urllib.parse import urljoin, urlparse, urlunparse, ParseResult

import requests
from PIL import Image

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.common.exceptions import (
    TimeoutException,
    StaleElementReferenceException,
    WebDriverException,
)
from selenium.webdriver.support.ui import WebDriverWait

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

import urllib.robotparser as robotparser


# ----------------------------- Utility helpers -----------------------------

def wait_for_ready(driver, timeout=25):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )

def smooth_scroll_to_bottom(driver, step_px=1000, max_wait_s=6):
    """Trigger lazy content with a gentle scroll."""
    try:
        last_height = driver.execute_script("return document.body.scrollHeight")
    except Exception:
        return
    start = time.time()
    while True:
        try:
            driver.execute_script(f"window.scrollBy(0, {step_px});")
        except Exception:
            break
        time.sleep(0.35)
        try:
            new_height = driver.execute_script("return document.body.scrollHeight")
        except Exception:
            break
        if new_height == last_height or (time.time() - start) > max_wait_s:
            break
        last_height = new_height

def sanitize_text(s: str, max_len=250):
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s).strip()
    return s[:max_len]

def is_useful_href(href: str, include_fragments=False):
    if not href:
        return False
    href = href.strip()
    if href.startswith(("javascript:", "mailto:", "tel:", "sms:", "data:")):
        return False
    if not include_fragments and href.startswith("#"):
        return False
    return True

def highlight_element(driver, element):
    try:
        driver.execute_script(
            "arguments[0].setAttribute('data-prev-style', arguments[0].getAttribute('style') || '');"
            "arguments[0].style.outline='3px solid #ff0066';"
            "arguments[0].style.outlineOffset='2px';",
            element,
        )
    except Exception:
        pass

def unhighlight_element(driver, element):
    try:
        driver.execute_script(
            "var prev = arguments[0].getAttribute('data-prev-style');"
            "if (prev !== null) { arguments[0].setAttribute('style', prev); }"
            "arguments[0].removeAttribute('data-prev-style');",
            element,
        )
    except Exception:
        pass

def element_screenshot(element, out_path):
    """Cropped screenshot of the WebElement (Selenium 4)."""
    element.screenshot(out_path)
    return out_path

def resize_image_to_fit_save(in_path, out_path, max_w=380, max_h=140):
    with Image.open(in_path) as im:
        im = im.convert("RGBA")
        w, h = im.size
        scale = min(max_w / w, max_h / h, 1.0)
        if scale < 1.0:
            new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
            im = im.resize(new_size, Image.Resampling.LANCZOS)
        im.save(out_path, format="PNG")
        return im.size  # (w, h) after resize

def px_to_row_height_points(px):
    # Excel row height unit ~ points; approx 1px ≈ 0.75pt
    return max(18, math.ceil(px * 0.75))

def http_resolve(url, timeout=8):
    """Resolve final URL + status (HEAD with GET fallback)."""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; LinkAuditBot/1.0)"}
        r = requests.head(url, allow_redirects=True, timeout=timeout, headers=headers)
        if r.status_code in (405, 403, 400):
            r = requests.get(url, allow_redirects=True, timeout=timeout, headers=headers, stream=True)
        return (r.url, r.status_code)
    except Exception:
        return (url, None)

def excel_set_col_widths(ws, widths):
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_workbook():
    wb = Workbook()
    ws_links = wb.active
    ws_links.title = "Links"
    ws_links.append([
        "Source Page",
        "Source Title",
        "Screenshot",
        "Link Text",
        "Raw Href (absolute)",
        "Resolved URL",
        "HTTP Status",
        "Target",
        "Rel",
        "Internal/External",
    ])
    excel_set_col_widths(ws_links, [50, 40, 55, 45, 80, 80, 12, 12, 18, 16])

    ws_summary = wb.create_sheet("Crawl Summary")
    ws_summary.append(["#","Page URL","Page Title","HTTP Status (fetch)","Links Captured"])
    excel_set_col_widths(ws_summary, [6, 90, 45, 20, 20])
    return wb, ws_links, ws_summary

def normalize_url(base_url, href, keep_query=False):
    """Return absolute normalized URL (strip fragment; optional query)."""
    abs_url = urljoin(base_url, href)
    parsed = urlparse(abs_url)
    if parsed.scheme not in ("http", "https"):
        return None
    # Remove fragment
    parsed = ParseResult(parsed.scheme, parsed.netloc, parsed.path, parsed.params,
                         parsed.query if keep_query else "", "")
    return urlunparse(parsed)

def host_core(host: str):
    """Remove leading 'www.' for a broader same-domain match (approx)."""
    host = (host or "").lower()
    return host[4:] if host.startswith("www.") else host

def is_internal_url(url: str, base_host: str, include_subdomains=True):
    """Heuristic internal check without extra deps (treat base host and its subdomains as internal)."""
    try:
        netloc = urlparse(url).netloc.lower()
    except Exception:
        return False
    base_core = host_core(base_host)
    netloc_core = host_core(netloc)
    if netloc_core == base_core:
        return True
    if include_subdomains and netloc_core.endswith("." + base_core):
        return True
    return False

def allowed_by_robots(robots, url, user_agent="*"):
    try:
        return robots.can_fetch(user_agent, url)
    except Exception:
        return True
# ------------------------------- Main logic --------------------------------

def crawl_and_report(args):
    start_url = args.start_url
    base_host = urlparse(start_url).netloc

    # Prepare Selenium Chrome
    chrome_opts = ChromeOptions()
    if not args.headful:
        chrome_opts.add_argument("--headless=new")
    chrome_opts.add_argument("--window-size=1400,2000")
    chrome_opts.add_argument("--disable-gpu")
    chrome_opts.add_argument("--no-sandbox")
    chrome_opts.add_argument("--disable-dev-shm-usage")
    chrome_opts.add_argument("--disable-blink-features=AutomationControlled")
    chrome_opts.add_argument("--force-device-scale-factor=1")
    chrome_opts.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    )

    try:
        driver = webdriver.Chrome(options=chrome_opts)
    except WebDriverException:
        print("Failed to start Chrome WebDriver. Ensure Chrome is installed and Selenium >= 4.6.")
        raise

    # robots.txt
    robots = robotparser.RobotFileParser()
    robots_read = False
    if args.respect_robots:
        try:
            origin = urlparse(start_url)
            robots_url = f"{origin.scheme}://{origin.netloc}/robots.txt"
            robots.set_url(robots_url)
            robots.read()
            robots_read = True
        except Exception:
            robots_read = False

    temp_dir = tempfile.mkdtemp(prefix="link_crawl_screens_")

    # Excel workbook
    wb, ws_links, ws_summary = create_workbook()

    # BFS queue and visited
    q = deque([(start_url, 0)])
    visited = set()
    pages_crawled = 0
    total_links_captured = 0
    link_img_max_w, link_img_max_h = 380, 140

    if args.pause_on_first:
        # Open the first page headfully, so user can login, then continue
        if not args.headful:
            print("NOTE: --pause-on-first is most useful with --headful")
        print("Opening the start URL to let you login/solve CAPTCHA if needed…")
        driver.get(start_url)
        try:
            wait_for_ready(driver, timeout=40)
            input("Press ENTER to start the crawl…")
        except Exception:
            pass
        # Put it back on the queue for normal processing
        q = deque([(driver.current_url, 0)])

    while q and pages_crawled < args.max_pages:
        page_url, depth = q.popleft()

        # Respect robots
        if args.respect_robots and robots_read:
            if not allowed_by_robots(robots, page_url):
                # Skip page disallowed by robots.txt
                continue

        # Normalize page URL for visited set
        normalized_page = normalize_url(page_url, "", keep_query=args.keep_query)
        if not normalized_page:
            continue
        if normalized_page in visited:
            continue

        visited.add(normalized_page)
        pages_crawled += 1

        # Fetch page
        try:
            driver.get(page_url)
            wait_for_ready(driver, timeout=35)
            smooth_scroll_to_bottom(driver, step_px=1200, max_wait_s=7)
            page_title = sanitize_text(driver.title, 300)
            # Infer fetch status via a lightweight HEAD (optional)
            fetch_status = None
            if not args.no_resolve:
                _, fetch_status = http_resolve(driver.current_url, timeout=args.timeout)
        except TimeoutException:
            page_title = ""
            fetch_status = None
        except Exception:
            page_title = ""
            fetch_status = None

        # Collect links on this page
        try:
            anchors = driver.find_elements(By.TAG_NAME, "a")
        except Exception:
            anchors = []

        # First, collect link data and also discover next URLs for BFS
        rows_for_page = []
        discovered_urls = set()

        for idx, a in enumerate(anchors):
            # Skip stale/hidden/zero-size quickly
            try:
                href = a.get_attribute("href")
            except StaleElementReferenceException:
                continue
            except Exception:
                continue

            if not is_useful_href(href, include_fragments=args.include_fragments):
                continue

            abs_url = normalize_url(driver.current_url, href, keep_query=args.keep_query)
            if not abs_url:
                continue

            # BFS discovery (only from links that look internal or allowed by filters)
            consider_for_crawl = True
            if args.same_domain_only and not is_internal_url(abs_url, base_host, include_subdomains=args.include_subdomains):
                consider_for_crawl = False

            if args.pattern_include and not re.search(args.pattern_include, abs_url, flags=re.IGNORECASE):
                consider_for_crawl = False
            if args.pattern_exclude and re.search(args.pattern_exclude, abs_url, flags=re.IGNORECASE):
                consider_for_crawl = False

            if consider_for_crawl and depth < args.depth:
                discovered_urls.add(abs_url)

            # Metadata for Excel
            try:
                link_text = a.text or a.get_attribute("aria-label") or ""
                link_text = sanitize_text(link_text, max_len=250)
                target = a.get_attribute("target") or ""
                rel = a.get_attribute("rel") or ""

                # Prepare screenshot if enabled
                shot_path = ""
                resized_path = ""
                row_height_px = 60

                if not args.no_screenshots:
                    # Only screenshot visible elements with some size
                    try:
                        if a.is_displayed():
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", a)
                            time.sleep(0.12)
                            highlight_element(driver, a)
                            time.sleep(0.05)

                            shot_path = os.path.join(temp_dir, f"p{pages_crawled}_link_{idx+1}.png")
                            element_screenshot(a, shot_path)

                            unhighlight_element(driver, a)

                            # Resize to fit the Excel cell nicely
                            resized_path = os.path.join(temp_dir, f"p{pages_crawled}_link_{idx+1}_res.png")
                            (rw, rh) = resize_image_to_fit_save(
                                shot_path, resized_path, max_w=link_img_max_w, max_h=link_img_max_h
                            )
                            row_height_px = rh if rh else 60
                        else:
                            # not displayed; skip screenshot
                            resized_path = ""
                    except Exception:
                        resized_path = ""
                # Resolve final URL / status
                final_url, status = (abs_url, None)
                if not args.no_resolve:
                    final_url, status = http_resolve(abs_url, timeout=args.timeout)

                internal_external = "Internal" if is_internal_url(final_url or abs_url, base_host, include_subdomains=args.include_subdomains) else "External"

                rows_for_page.append({
                    "source_url": driver.current_url,
                    "source_title": page_title,
                    "img_path": resized_path,  # may be empty
                    "link_text": link_text,
                    "abs_url": abs_url,
                    "final_url": final_url,
                    "status": status,
                    "target": target,
                    "rel": rel,
                    "scope": internal_external,
                    "row_height_px": row_height_px
                })

                if len(rows_for_page) >= args.max_links_per_page:
                    break

            except StaleElementReferenceException:
                continue
            except Exception:
                continue

        # Add discovered URLs to queue
        for u in discovered_urls:
            if args.respect_robots and robots_read and not allowed_by_robots(robots, u):
                continue
            norm_u = normalize_url(u, "", keep_query=args.keep_query)
            if norm_u and norm_u not in visited:
                q.append((u, depth + 1))

        # Write summary row
        ws_summary.append([pages_crawled, driver.current_url, page_title, fetch_status, len(rows_for_page)])

        # Write link rows to Excel
        start_row = ws_links.max_row + 1
        for i, rec in enumerate(rows_for_page, start=0):
            row = start_row + i
            # Text columns
            ws_links.cell(row=row, column=1, value=rec["source_url"])
            ws_links.cell(row=row, column=2, value=rec["source_title"])

            # Image anchor in column C (if available)
            if rec["img_path"]:
                try:
                    xl_img = XLImage(rec["img_path"])
                    ws_links.add_image(xl_img, f"C{row}")
                    ws_links.row_dimensions[row].height = px_to_row_height_points(rec["row_height_px"])
                except Exception:
                    ws_links.row_dimensions[row].height = 60
            else:
                ws_links.row_dimensions[row].height = 60

            ws_links.cell(row=row, column=4, value=rec["link_text"])
            ws_links.cell(row=row, column=5, value=rec["abs_url"])
            ws_links.cell(row=row, column=6, value=rec["final_url"])
            ws_links.cell(row=row, column=7, value=rec["status"] if rec["status"] is not None else "")
            ws_links.cell(row=row, column=8, value=rec["target"])
            ws_links.cell(row=row, column=9, value=rec["rel"])
            ws_links.cell(row=row, column=10, value=rec["scope"])

        total_links_captured += len(rows_for_page)

        # Politeness delay
        time.sleep(max(0.0, args.delay))

        if total_links_captured >= args.max_total_links:
            break
    # Save workbook
    wb.save(args.out)
    print(f"✅ Crawl complete: {pages_crawled} page(s) visited")
    print(f"✅ Links captured: {total_links_captured}")
    print(f"📄 Report saved: {os.path.abspath(args.out)}")

    # Cleanup
    driver.quit()
    shutil.rmtree(temp_dir, ignore_errors=True)


def parse_args():
    p = argparse.ArgumentParser(description="Crawl site pages and export link report to Excel with screenshots.")
    p.add_argument("--start-url", required=True, help="Start URL for the crawl")
    p.add_argument("--out", default="site_links_report.xlsx", help="Output Excel path")

    # Crawl controls
    p.add_argument("--max-pages", type=int, default=30, help="Max number of pages to crawl")
    p.add_argument("--max-links-per-page", type=int, default=300, help="Max links per page to capture")
    p.add_argument("--max-total-links", type=int, default=3000, help="Global max links to capture")
    p.add_argument("--depth", type=int, default=2, help="Max crawl depth from start URL")
    p.add_argument("--delay", type=float, default=0.5, help="Delay (seconds) between page fetches")

    p.add_argument("--same-domain-only", action="store_true", default=True,
                   help="Restrict crawl to same domain (default: on)")
    p.add_argument("--include-subdomains", action="store_true", help="Treat subdomains as internal")
    p.add_argument("--keep-query", action="store_true", help="Do not strip query params during normalization")
    p.add_argument("--include-fragments", action="store_true", help="Include #fragment-only links on pages")
    p.add_argument("--pattern-include", default="", help="Regex; only URLs matching are crawled")
    p.add_argument("--pattern-exclude", default="", help="Regex; URLs matching are excluded from crawl and capture")

    # Networking / robots
    p.add_argument("--timeout", type=int, default=8, help="HTTP timeout seconds for resolution")
    group_robots = p.add_mutually_exclusive_group()
    group_robots.add_argument("--respect-robots", action="store_true", help="Obey robots.txt (recommended)")
    group_robots.add_argument("--ignore-robots", action="store_true", help="Ignore robots.txt (use responsibly)")

    # Browser
    p.add_argument("--headful", action="store_true", help="Run a visible browser")
    p.add_argument("--pause-on-first", action="store_true", help="Open first page and wait for ENTER (login etc.)")

    # Output speed toggles
    p.add_argument("--no-resolve", action="store_true", help="Skip resolving final URLs / status (faster)")
    p.add_argument("--no-screenshots", action="store_true", help="Do not capture element screenshots")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    crawl_and_report(args)
